"""Material de conferencia: PAGAMENTOS EFETIVOS de Consultoria e Cobranca do
fundo REALINVEST FIDC (visao de caixa), sem a provisao.

Companion do export_conferencia_consultoria_cobranca_xlsx.py (que cruza
provisao x pagamento). Aqui o escopo e SO o pagamento identificado no caixa.

Gera um .xlsx com 4 abas:
  - Leia-me            : escopo, fonte, convencao de sinal, nota do estorno
  - Pagamentos         : 1 linha por lancamento de caixa (saida/estorno/liquido)
  - Resumo mensal      : por mes de liquidacao (MM/AAAA) x categoria, pago liquido
  - Resumo categoria   : total bruto (saidas) + estornos + liquido por categoria

CONVENCAO DE SINAL (importante): valores ficam COM SINAL.
  saidas   = negativo (dinheiro saindo)
  entradas = positivo (estorno / reclassificacao)
  liquido  = entradas + saidas  (a metrica de verdade)
Somar a coluna 'Liquido' anula automaticamente os pares lancamento+estorno
(ex.: 29/08 e 31/10/2025: -100k 'Despesa de Consultoria Especializada'
estornado por +100k 'Taxa de Consultoria Especializada' = wash). NAO somar
magnitudes (abs), senao o wash infla o total.

Fonte: wh_movimento_caixa (visao de caixa QiTech), conta da carteira REALINVEST.
Le DATABASE_URL de backend/.env.
"""

from __future__ import annotations

import asyncio
import re
from datetime import date
from decimal import Decimal
from pathlib import Path

import asyncpg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BACKEND_DIR = Path(__file__).resolve().parent.parent
ENV_PATH = BACKEND_DIR / ".env"
OUT_PATH = Path(r"C:\app_gr\pagamentos_consultoria_cobranca.xlsx")

_MONEY = "#,##0.00"
_DATE = "DD/MM/YYYY"
_HEAD_FILL = PatternFill("solid", fgColor="1F2937")
_HEAD_FONT = Font(bold=True, color="FFFFFF")


def _load_dsn() -> str:
    text = ENV_PATH.read_text(encoding="utf-8")
    m = re.search(r"^\s*DATABASE_URL\s*=\s*(.+)\s*$", text, re.MULTILINE)
    if not m:
        raise RuntimeError("DATABASE_URL not found in .env")
    url = m.group(1).strip().strip('"').strip("'")
    return url.replace("postgresql+asyncpg://", "postgresql://")


_CAT = (
    "CASE WHEN descricao ILIKE '%cobran%' OR historico_traduzido ILIKE '%cobran%' "
    "THEN 'Cobranca' ELSE 'Consultoria' END"
)
_FILTER = (
    "descricao ILIKE '%consultor%' OR historico_traduzido ILIKE '%consultor%' "
    "OR descricao ILIKE '%cobran%' OR historico_traduzido ILIKE '%cobran%'"
)
_SCOPE = "carteira_cliente_nome='REALINVEST FIDC'"

SQL_PAGAMENTOS = f"""
SELECT data_liquidacao AS data,
       {_CAT} AS categoria,
       descricao, historico_traduzido,
       saidas, entradas, (entradas + saidas) AS liquido
FROM wh_movimento_caixa
WHERE {_SCOPE} AND ({_FILTER})
ORDER BY data_liquidacao, categoria, descricao;
"""

SQL_RESUMO_MES = f"""
SELECT to_char(data_liquidacao,'YYYY-MM') AS mes,
       {_CAT} AS categoria,
       sum(entradas + saidas) AS pago_liquido,
       count(*) AS lancamentos
FROM wh_movimento_caixa
WHERE {_SCOPE} AND ({_FILTER})
GROUP BY 1, 2
ORDER BY 1, 2;
"""

SQL_RESUMO_CAT = f"""
SELECT {_CAT} AS categoria,
       sum(saidas) AS saidas_brutas,
       sum(entradas) AS estornos,
       sum(entradas + saidas) AS liquido,
       count(*) AS lancamentos
FROM wh_movimento_caixa
WHERE {_SCOPE} AND ({_FILTER})
GROUP BY 1
ORDER BY 1;
"""


def _f(v: object) -> object:
    return float(v) if isinstance(v, Decimal) else v


def _header(ws, labels: list[str]) -> None:
    ws.append(labels)
    for c in ws[1]:
        c.fill = _HEAD_FILL
        c.font = _HEAD_FONT
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _total_row(ws, *, label_col: int, label: str, value_cols: dict[int, float]) -> None:
    row = ws.max_row + 2
    ws.cell(row, label_col, label).font = Font(bold=True)
    for col, val in value_cols.items():
        c = ws.cell(row, col, val)
        c.font = Font(bold=True)
        c.number_format = _MONEY


def _build_obs(pags: list) -> dict[int, str]:
    """Flag estorno rows and their same-day/same-categoria paired debits (wash)."""
    estornos: dict[tuple, list[float]] = {}
    for r in pags:
        ent = _f(r["entradas"]) or 0.0
        if ent > 0:
            estornos.setdefault((r["data"], r["categoria"]), []).append(round(ent, 2))
    obs: dict[int, str] = {}
    for i, r in enumerate(pags):
        ent = _f(r["entradas"]) or 0.0
        sai = _f(r["saidas"]) or 0.0
        if ent > 0:
            obs[i] = "estorno (reclassificacao)"
        elif round(abs(sai), 2) in estornos.get((r["data"], r["categoria"]), []):
            obs[i] = "lancamento estornado (anula)"
        else:
            obs[i] = ""
    return obs


async def main() -> None:
    conn = await asyncpg.connect(_load_dsn())
    try:
        pags = await conn.fetch(SQL_PAGAMENTOS)
        res_mes = await conn.fetch(SQL_RESUMO_MES)
        res_cat = await conn.fetch(SQL_RESUMO_CAT)
    finally:
        await conn.close()

    obs = _build_obs(pags)
    wb = Workbook()

    # ── Leia-me ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Leia-me"
    notas = [
        ("Pagamentos efetivos — Consultoria e Cobranca", True),
        ("Fundo: REALINVEST FIDC  |  Gerado em: " + date.today().strftime("%d/%m/%Y"), False),
        ("", False),
        ("Escopo: SO o pagamento efetivo (visao de caixa). A provisao (CPR) esta", False),
        ("na outra planilha (conferencia_consultoria_cobranca.xlsx).", False),
        ("", False),
        ("Fonte: wh_movimento_caixa (visao de caixa QiTech), carteira REALINVEST FIDC.", False),
        ("", False),
        ("Convencao de SINAL (valores com sinal — NAO usar magnitude):", True),
        ("  Saida    = negativo (dinheiro saindo do fundo)", False),
        ("  Estorno  = positivo (entradas; reclassificacao / estorno de lancamento)", False),
        ("  Liquido  = Estorno + Saida  -> a metrica de verdade do pago", False),
        ("", False),
        ("ATENCAO (pares lancamento+estorno = wash):", True),
        ("  Em 29/08 e 31/10/2025 ha -100.000 'Despesa de Consultoria Especializada'", False),
        ("  estornado por +100.000 'Taxa de Consultoria Especializada' no MESMO dia.", False),
        ("  Esses pares se ANULAM. Por isso o total usa a coluna Liquido (com sinal):", False),
        ("  somar magnitudes (abs) inflaria o bruto em 200k que nunca saiu de fato.", False),
        ("  O pagamento real do mes nesses casos e a linha 'Consultoria' de -50.000.", False),
        ("", False),
        ("Classificacao: historico/descricao com 'cobran' -> Cobranca; senao Consultoria.", False),
    ]
    for txt, bold in notas:
        ws.append([txt])
        if bold:
            ws[ws.max_row][0].font = Font(bold=True)
    ws.column_dimensions["A"].width = 92

    # ── Pagamentos ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Pagamentos")
    _header(ws, ["Data", "Categoria", "Rubrica (descricao)", "Historico traduzido",
                 "Saida", "Estorno", "Liquido", "Obs"])
    for i, r in enumerate(pags):
        ws.append([r["data"], r["categoria"], r["descricao"], r["historico_traduzido"],
                   _f(r["saidas"]), _f(r["entradas"]), _f(r["liquido"]), obs[i]])
        row = ws.max_row
        ws.cell(row, 1).number_format = _DATE
        for col in (5, 6, 7):
            ws.cell(row, col).number_format = _MONEY
    _total_row(ws, label_col=1, label="TOTAL (liquido)", value_cols={
        5: sum(_f(r["saidas"]) for r in pags),
        6: sum(_f(r["entradas"]) for r in pags),
        7: sum(_f(r["liquido"]) for r in pags),
    })
    _widths(ws, [12, 12, 38, 30, 15, 15, 15, 28])

    # ── Resumo mensal ────────────────────────────────────────────────────
    ws = wb.create_sheet("Resumo mensal")
    _header(ws, ["Mes (MM/AAAA)", "Categoria", "Pago liquido", "Lancamentos"])
    for r in res_mes:
        mes = r["mes"]
        ws.append([f"{mes[5:7]}/{mes[0:4]}", r["categoria"], _f(r["pago_liquido"]), r["lancamentos"]])
        ws.cell(ws.max_row, 3).number_format = _MONEY
    _total_row(ws, label_col=1, label="TOTAL", value_cols={
        3: sum(_f(r["pago_liquido"]) for r in res_mes),
    })
    _widths(ws, [16, 12, 16, 14])

    # ── Resumo por categoria ─────────────────────────────────────────────
    ws = wb.create_sheet("Resumo categoria")
    _header(ws, ["Categoria", "Saidas brutas", "Estornos", "Liquido", "Lancamentos"])
    for r in res_cat:
        ws.append([r["categoria"], _f(r["saidas_brutas"]), _f(r["estornos"]),
                   _f(r["liquido"]), r["lancamentos"]])
        row = ws.max_row
        for col in (2, 3, 4):
            ws.cell(row, col).number_format = _MONEY
    _total_row(ws, label_col=1, label="TOTAL", value_cols={
        2: sum(_f(r["saidas_brutas"]) for r in res_cat),
        3: sum(_f(r["estornos"]) for r in res_cat),
        4: sum(_f(r["liquido"]) for r in res_cat),
    })
    _widths(ws, [14, 16, 16, 16, 14])

    wb.save(OUT_PATH)
    print(f"OK -> {OUT_PATH}")
    print(f"  Pagamentos: {len(pags)} | Resumo mensal: {len(res_mes)} | Categorias: {len(res_cat)}")


if __name__ == "__main__":
    asyncio.run(main())
