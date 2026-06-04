"""One-off: export all rows of wh_cpr_movimento to an .xlsx file.

Reads DATABASE_URL from backend/.env, connects via asyncpg, dumps every
row (ordered) into a single worksheet. UUID -> str, Decimal -> float,
date/datetime kept native so Excel treats them as real dates.
"""

from __future__ import annotations

import asyncio
import os
import re
from datetime import datetime
from decimal import Decimal
from pathlib import Path
from uuid import UUID

import asyncpg
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

BACKEND_DIR = Path(__file__).resolve().parent.parent
ENV_PATH = BACKEND_DIR / ".env"
OUT_PATH = Path(r"C:\app_gr\wh_cpr_movimento.xlsx")
TABLE = "wh_cpr_movimento"


def _load_dsn() -> str:
    """Parse DATABASE_URL from .env and convert to a plain asyncpg DSN."""
    text = ENV_PATH.read_text(encoding="utf-8")
    m = re.search(r"^\s*DATABASE_URL\s*=\s*(.+)\s*$", text, re.MULTILINE)
    if not m:
        raise RuntimeError("DATABASE_URL not found in .env")
    url = m.group(1).strip().strip('"').strip("'")
    # asyncpg wants 'postgresql://', not SQLAlchemy's '+asyncpg' dialect tag.
    return url.replace("postgresql+asyncpg://", "postgresql://")


def _cell(value: object) -> object:
    if isinstance(value, UUID):
        return str(value)
    if isinstance(value, Decimal):
        return float(value)
    if isinstance(value, datetime) and value.tzinfo is not None:
        # Excel can't store tz-aware datetimes; drop tzinfo (UTC from PG).
        return value.replace(tzinfo=None)
    return value


async def main() -> None:
    dsn = _load_dsn()
    conn = await asyncpg.connect(dsn)
    try:
        rows = await conn.fetch(
            f"SELECT * FROM {TABLE} ORDER BY data_posicao, carteira_cliente_doc, id"
        )
    finally:
        await conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "wh_cpr_movimento"

    if not rows:
        ws["A1"] = "(tabela vazia)"
        wb.save(OUT_PATH)
        print(f"0 rows -> {OUT_PATH}")
        return

    columns = list(rows[0].keys())
    ws.append(columns)
    for c in ws[1]:
        c.font = Font(bold=True)
    ws.freeze_panes = "A2"

    for r in rows:
        ws.append([_cell(r[c]) for c in columns])

    # Reasonable column widths (cap at 40).
    for idx, col in enumerate(columns, start=1):
        width = min(max(len(col) + 2, 12), 40)
        ws.column_dimensions[get_column_letter(idx)].width = width

    ws.auto_filter.ref = f"A1:{get_column_letter(len(columns))}{len(rows) + 1}"
    wb.save(OUT_PATH)
    print(f"{len(rows)} rows x {len(columns)} cols -> {OUT_PATH}")


if __name__ == "__main__":
    asyncio.run(main())
