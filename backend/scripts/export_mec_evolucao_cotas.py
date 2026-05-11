"""Export wh_mec_evolucao_cotas (silver MEC) para Excel.

Le todas as linhas da silver MEC (apenas colunas de negocio) e grava num
.xlsx com header em negrito, freeze pane na primeira linha e formato
numerico apropriado por tipo de coluna (moeda / cota / variacao %).

Read-only. Nao registra decision_log (export tecnico, nao e decisao de
dominio -- ver CLAUDE.md sec 14.2).

Uso (de backend/, com .venv ativo):
    .venv\\Scripts\\python.exe scripts/export_mec_evolucao_cotas.py
    .venv\\Scripts\\python.exe scripts/export_mec_evolucao_cotas.py --tenant a7-credit
    .venv\\Scripts\\python.exe scripts/export_mec_evolucao_cotas.py --out C:/tmp/mec.xlsx

Args:
    --tenant <uuid|slug>      filtra um tenant (default: todos)
    --out <path>              caminho do .xlsx (default: scripts/out/mec_evolucao_cotas_<ts>.xlsx)
"""

from __future__ import annotations

import argparse
import asyncio
import sys
from datetime import UTC, datetime
from decimal import Decimal
from pathlib import Path
from uuid import UUID

# Side-effect imports (registry SQLAlchemy completo).
import app.shared.identity.tenant  # noqa: F401
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from sqlalchemy import text

from app.core.database import AsyncSessionLocal

# Ordem das colunas no xlsx + formato numerico por coluna. "@" = texto.
COLUMNS: list[tuple[str, str]] = [
    ("tenant", "@"),
    ("unidade_administrativa", "@"),
    ("data_posicao", "yyyy-mm-dd"),
    ("carteira_cliente_id", "@"),
    ("carteira_cliente_nome", "@"),
    ("carteira_cliente_doc", "@"),
    ("entradas", "#,##0.00"),
    ("saidas", "#,##0.00"),
    ("aporte", "#,##0.00"),
    ("retirada", "#,##0.00"),
    ("patrimonio", "#,##0.00"),
    ("quantidade", "#,##0.00000000"),
    ("valor_da_cota", "#,##0.00000000"),
    ("variacao_diaria", "0.0000"),
    ("variacao_mensal", "0.0000"),
    ("variacao_anual", "0.0000"),
    ("variacao_total", "0.0000"),
]

_BASE_QUERY = """
    SELECT
        t.slug                 AS tenant,
        ua.nome                AS unidade_administrativa,
        m.data_posicao,
        m.carteira_cliente_id,
        m.carteira_cliente_nome,
        m.carteira_cliente_doc,
        m.entradas, m.saidas, m.aporte, m.retirada,
        m.patrimonio, m.quantidade, m.valor_da_cota,
        m.variacao_diaria, m.variacao_mensal,
        m.variacao_anual, m.variacao_total
      FROM wh_mec_evolucao_cotas m
      JOIN tenants t ON t.id = m.tenant_id
      LEFT JOIN cadastros_unidade_administrativa ua
             ON ua.id = m.unidade_administrativa_id
"""

_ORDER_BY = " ORDER BY t.slug, m.carteira_cliente_nome, m.data_posicao"


def _build_query(
    tenant_uuid: UUID | None, tenant_slug: str | None
) -> tuple[object, dict[str, object]]:
    where_parts: list[str] = []
    params: dict[str, object] = {}
    if tenant_uuid is not None:
        where_parts.append("m.tenant_id = :tenant_uuid")
        params["tenant_uuid"] = str(tenant_uuid)
    if tenant_slug is not None:
        where_parts.append("t.slug = :tenant_slug")
        params["tenant_slug"] = tenant_slug
    where_clause = (" WHERE " + " AND ".join(where_parts)) if where_parts else ""
    return text(_BASE_QUERY + where_clause + _ORDER_BY), params


def _parse_tenant(raw: str | None) -> tuple[UUID | None, str | None]:
    """Devolve (uuid, slug). Exatamente um dos dois e nao-None se raw vier."""
    if not raw:
        return None, None
    try:
        return UUID(raw), None
    except ValueError:
        return None, raw


def _default_out_path() -> Path:
    ts = datetime.now(UTC).strftime("%Y%m%d-%H%M%S")
    return Path(__file__).resolve().parent / "out" / f"mec_evolucao_cotas_{ts}.xlsx"


async def _fetch_rows(tenant_uuid: UUID | None, tenant_slug: str | None) -> list[dict]:
    stmt, params = _build_query(tenant_uuid, tenant_slug)
    async with AsyncSessionLocal() as db:
        result = await db.execute(stmt, params)
        return [dict(row) for row in result.mappings().all()]


def _to_excel_value(raw: object) -> object:
    if isinstance(raw, Decimal):
        return float(raw)
    return raw


def _write_workbook(rows: list[dict], out_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "MEC"

    header_font = Font(bold=True)
    header_align = Alignment(horizontal="left", vertical="center")
    for col_idx, (name, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.font = header_font
        cell.alignment = header_align

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (name, fmt) in enumerate(COLUMNS, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=_to_excel_value(row[name]))
            cell.number_format = fmt

    ws.freeze_panes = "A2"

    for col_idx, (name, _) in enumerate(COLUMNS, start=1):
        max_len = len(name)
        for row in rows:
            value = row[name]
            if value is None:
                continue
            text_repr = (
                value.isoformat() if hasattr(value, "isoformat") else str(value)
            )
            if len(text_repr) > max_len:
                max_len = len(text_repr)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


async def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--tenant", help="UUID ou slug do tenant; default = todos")
    parser.add_argument("--out", type=Path, help="path do .xlsx de saida")
    args = parser.parse_args()

    tenant_uuid, tenant_slug = _parse_tenant(args.tenant)
    out_path: Path = args.out or _default_out_path()

    started = datetime.now(UTC)
    rows = await _fetch_rows(tenant_uuid, tenant_slug)
    elapsed_query = (datetime.now(UTC) - started).total_seconds()

    if not rows:
        print("[export-mec] 0 linhas retornadas -- nada a exportar.", file=sys.stderr)
        return 1

    _write_workbook(rows, out_path)
    size_kb = out_path.stat().st_size / 1024
    print(
        f"[export-mec] {len(rows)} linhas lidas em {elapsed_query:.2f}s",
    )
    print(f"[export-mec] arquivo: {out_path} ({size_kb:.0f} KB)")
    return 0


if __name__ == "__main__":
    sys.exit(asyncio.run(main()))
