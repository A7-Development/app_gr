"""Pagamentos de CONSULTORIA / COBRANCA do REALINVEST FIDC, por EMPRESA.

Substitui o recorte do export_pagamentos_consultoria_cobranca_xlsx.py (que era
caixa-based e SO via a conta conciliacao -> incompleto, faltava abril). Aqui a
fonte e o EXTRATO BANCARIO real (`wh_extrato_bancario`), as DUAS contas QiTech
(conciliacao 4532551 + movimento 4532543), com a empresa (contraparte) nominal.

Universo = TEDs a debito para os prestadores de servico confirmados:
  - ONBOARD CONSULTORIA LTDA   (45934845000192)  consultoria + cobranca, 03/2025->
  - BLUESTONE CONS. E ASSESS.  (11314857000100)  consultoria, 03/2022->03/2025
Handoff de prestador em mar/2025 (BLUESTONE -> ONBOARD).

Como o universo foi descoberto (2026-05-27): cruzando os pagamentos que a visao
de caixa rotula como consultoria/cobranca contra os TEDs do extrato (mesmo valor,
data +/-7d) -> revela a contraparte; + varredura por nome
(consultor/assessor/cobran/recupera/gestao) no extrato. VALOREN RECUPERADORA DE
RESIDUOS (23571153000194) bateu no keyword mas e CEDENTE (reciclagem), nao
servico -> EXCLUIDA. Reabrir a lista se surgir prestador novo.

Cross-check: total extrato (R$ 1.620.006,39) = total caixa (R$ 1.520.006,39) +
R$ 100k do pagamento de abril que a visao de caixa nao bookou.

NB: a categoria consultoria-vs-cobranca NAO e separavel no extrato (ambas saem
como TED a ONBOARD; 2x 50k/mes = 1 consultoria + 1 cobranca). Por isso a coluna
'Tipo de servico' e por EMPRESA, nao por TED.

Le DATABASE_URL de backend/.env.
"""

from __future__ import annotations

import asyncio
import re
from datetime import date
from decimal import Decimal
from pathlib import Path

import asyncpg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BACKEND_DIR = Path(__file__).resolve().parent.parent
ENV_PATH = BACKEND_DIR / ".env"
OUT_PATH = Path(r"C:\app_gr\pagamentos_consultoria_cobranca_por_empresa.xlsx")

_MONEY = "#,##0.00"
_DATE = "DD/MM/YYYY"
_HEAD_FILL = PatternFill("solid", fgColor="1F2937")
_HEAD_FONT = Font(bold=True, color="FFFFFF")

# Prestadores confirmados: cnpj -> (nome canonico, tipo de servico)
_VENDORS: dict[str, tuple[str, str]] = {
    "45934845000192": ("ONBOARD CONSULTORIA LTDA", "Consultoria + Cobranca"),
    "11314857000100": ("BLUESTONE CONSULTORIA E ASSESSORIA FINANCEIRA LTDA", "Consultoria"),
}
_SHORT = {"45934845000192": "Onboard", "11314857000100": "Bluestone"}
_CONTA_LABEL = {"4532551": "Conciliacao", "4532543": "Movimento"}


def _load_dsn() -> str:
    text = ENV_PATH.read_text(encoding="utf-8")
    m = re.search(r"^\s*DATABASE_URL\s*=\s*(.+)\s*$", text, re.MULTILINE)
    if not m:
        raise RuntimeError("DATABASE_URL not found in .env")
    url = m.group(1).strip().strip('"').strip("'")
    return url.replace("postgresql+asyncpg://", "postgresql://")


SQL_PAGAMENTOS = """
SELECT COALESCE(data_movimento, data_lancamento) AS data,
       contrapartida_doc AS cnpj,
       conta,
       valor,
       descricao
FROM wh_extrato_bancario
WHERE tipo = 'D' AND contrapartida_doc = ANY($1::text[])
ORDER BY COALESCE(data_movimento, data_lancamento), conta;
"""


def _f(v: object) -> float:
    return float(v) if isinstance(v, Decimal) else float(v)


def _header(ws, labels: list[str]) -> None:
    ws.append(labels)
    for c in ws[1]:
        c.fill = _HEAD_FILL
        c.font = _HEAD_FONT
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _total_row(ws, *, label_col: int, label: str, value_cols: dict[int, float]) -> None:
    row = ws.max_row + 2
    ws.cell(row, label_col, label).font = Font(bold=True)
    for col, val in value_cols.items():
        c = ws.cell(row, col, val)
        c.font = Font(bold=True)
        c.number_format = _MONEY


async def main() -> None:
    conn = await asyncpg.connect(_load_dsn())
    try:
        rows = await conn.fetch(SQL_PAGAMENTOS, list(_VENDORS.keys()))
    finally:
        await conn.close()

    # Normaliza + agrega em Python.
    pays = []
    for r in rows:
        cnpj = r["cnpj"]
        nome, servico = _VENDORS[cnpj]
        pays.append({
            "data": r["data"],
            "nome": _SHORT.get(cnpj, nome),
            "empresa": nome,
            "cnpj": cnpj,
            "servico": servico,
            "conta": _CONTA_LABEL.get(r["conta"], r["conta"]),
            "valor": _f(r["valor"]),
            "desc": r["descricao"],
        })

    by_vendor: dict[str, dict] = {}
    by_vendor_month: dict[tuple[str, str], dict] = {}
    for p in pays:
        v = by_vendor.setdefault(p["cnpj"], {
            "nome": p["empresa"], "servico": p["servico"],
            "teds": 0, "total": 0.0, "min": p["data"], "max": p["data"],
        })
        v["teds"] += 1
        v["total"] += p["valor"]
        v["min"] = min(v["min"], p["data"])
        v["max"] = max(v["max"], p["data"])

        ym = p["data"].strftime("%Y-%m")
        m = by_vendor_month.setdefault((ym, p["cnpj"]), {
            "nome": p["empresa"], "teds": 0, "total": 0.0,
        })
        m["teds"] += 1
        m["total"] += p["valor"]

    wb = Workbook()

    # ── Leia-me ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Leia-me"
    notas = [
        ("Pagamentos de Consultoria / Cobranca por empresa", True),
        ("Fundo: REALINVEST FIDC  |  Gerado em: " + date.today().strftime("%d/%m/%Y"), False),
        ("", False),
        ("Fonte: wh_extrato_bancario (extrato bancario real) — fonte autoritativa.", False),
        ("Contas QiTech cobertas: conciliacao (0001/4532551) + movimento (0001/4532543).", False),
        ("Os pagamentos de servico saem da conta MOVIMENTO, com contraparte nominal.", False),
        ("", False),
        ("Prestadores confirmados (handoff em mar/2025):", True),
        ("  ONBOARD CONSULTORIA LTDA   (45934845000192) — consultoria + cobranca, 03/2025->", False),
        ("  BLUESTONE CONS. E ASSESS.  (11314857000100) — consultoria, 03/2022->03/2025", False),
        ("", False),
        ("Excluida: VALOREN RECUPERADORA DE RESIDUOS (23571153000194) — bateu no", False),
        ("  keyword 'recupera' mas e reciclagem/cedente, nao prestador de cobranca.", False),
        ("", False),
        ("NB consultoria vs cobranca: nao separavel no extrato (ambas = TED a ONBOARD;", False),
        ("  2x R$50k/mes = 1 consultoria + 1 cobranca). 'Tipo de servico' e por EMPRESA.", False),
        ("", False),
        ("Cross-check: total aqui (R$ 1.620.006,39) = total da visao de caixa", False),
        ("  (R$ 1.520.006,39) + R$ 100k do pagamento de abril/2026 que o caixa nao bookou.", False),
    ]
    for txt, bold in notas:
        ws.append([txt])
        if bold:
            ws[ws.max_row][0].font = Font(bold=True)
    ws.column_dimensions["A"].width = 94

    # ── Pagamentos ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Pagamentos")
    _header(ws, ["Data", "Nome", "Empresa", "CNPJ", "Tipo de servico", "Conta", "Valor", "Historico banco"])
    for p in pays:
        ws.append([p["data"], p["nome"], p["empresa"], p["cnpj"], p["servico"], p["conta"], p["valor"], p["desc"]])
        row = ws.max_row
        ws.cell(row, 1).number_format = _DATE
        ws.cell(row, 7).number_format = _MONEY
    _total_row(ws, label_col=1, label="TOTAL", value_cols={7: sum(p["valor"] for p in pays)})
    _widths(ws, [12, 11, 42, 16, 22, 14, 15, 46])

    # ── Resumo por empresa ───────────────────────────────────────────────
    ws = wb.create_sheet("Resumo por empresa")
    _header(ws, ["Empresa", "CNPJ", "Tipo de servico", "TEDs", "Total pago", "Primeiro", "Ultimo"])
    for cnpj, v in sorted(by_vendor.items(), key=lambda kv: -kv[1]["total"]):
        ws.append([v["nome"], cnpj, v["servico"], v["teds"], v["total"], v["min"], v["max"]])
        row = ws.max_row
        ws.cell(row, 5).number_format = _MONEY
        ws.cell(row, 6).number_format = _DATE
        ws.cell(row, 7).number_format = _DATE
    _total_row(ws, label_col=1, label="TOTAL", value_cols={5: sum(v["total"] for v in by_vendor.values())})
    _widths(ws, [46, 16, 22, 8, 16, 14, 14])

    # ── Resumo por empresa x mes ─────────────────────────────────────────
    ws = wb.create_sheet("Resumo empresa x mes")
    _header(ws, ["Mes (MM/AAAA)", "Empresa", "TEDs", "Total"])
    for (ym, _cnpj), m in sorted(by_vendor_month.items()):
        ws.append([f"{ym[5:7]}/{ym[0:4]}", m["nome"], m["teds"], m["total"]])
        ws.cell(ws.max_row, 4).number_format = _MONEY
    _total_row(ws, label_col=1, label="TOTAL", value_cols={
        4: sum(m["total"] for m in by_vendor_month.values()),
    })
    _widths(ws, [16, 46, 8, 16])

    wb.save(OUT_PATH)
    print(f"OK -> {OUT_PATH}")
    print(f"  Pagamentos: {len(pays)} | Empresas: {len(by_vendor)} | linhas empresa x mes: {len(by_vendor_month)}")
    for cnpj, v in sorted(by_vendor.items(), key=lambda kv: -kv[1]["total"]):
        print(f"    {v['nome'][:40]:40} {v['teds']:>3} TEDs  R$ {v['total']:>14,.2f}")
    print(f"    {'TOTAL':40} {sum(v['teds'] for v in by_vendor.values()):>3} TEDs  "
          f"R$ {sum(v['total'] for v in by_vendor.values()):>14,.2f}")


if __name__ == "__main__":
    asyncio.run(main())
