"""Material de conferencia: provisao (CPR) x pagamento (caixa) de Consultoria
e Cobranca do fundo REALINVEST FIDC.

Gera um .xlsx com 4 abas:
  - Leia-me        : escopo, fontes, convencoes (inclui a regra do acumulado)
  - Provisao (CPR) : 1 linha por lote, mes de competencia (MM/AAAA) + data de
                     pagamento agendada (DD/MM/AAAA) + valor PICO (saldo do
                     ultimo dia — NAO a soma dos snapshots diarios)
  - Pagamentos     : 1 linha por lancamento de caixa, data DD/MM/AAAA
  - Conciliacao    : por mes de pagamento (MM/AAAA), provisionado x pago x dif

IMPORTANTE: wh_cpr_movimento.valor e SALDO ACUMULADO da provisao (snapshot
diario que cresce ate o pagamento e zera). Por isso usamos o PICO por lote
(min(valor) = ultimo dia), nunca a soma.

Le DATABASE_URL de backend/.env (mesmo padrao do export_cpr_movimento_xlsx.py).
"""

from __future__ import annotations

import asyncio
import re
from datetime import date
from decimal import Decimal
from pathlib import Path

import asyncpg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BACKEND_DIR = Path(__file__).resolve().parent.parent
ENV_PATH = BACKEND_DIR / ".env"
OUT_PATH = Path(r"C:\app_gr\conferencia_consultoria_cobranca.xlsx")

_MONEY = "#,##0.00"
_DATE = "DD/MM/YYYY"
_HEAD_FILL = PatternFill("solid", fgColor="1F2937")
_HEAD_FONT = Font(bold=True, color="FFFFFF")


def _load_dsn() -> str:
    text = ENV_PATH.read_text(encoding="utf-8")
    m = re.search(r"^\s*DATABASE_URL\s*=\s*(.+)\s*$", text, re.MULTILINE)
    if not m:
        raise RuntimeError("DATABASE_URL not found in .env")
    url = m.group(1).strip().strip('"').strip("'")
    return url.replace("postgresql+asyncpg://", "postgresql://")


_CAT = (
    "CASE WHEN descricao ILIKE '%cobran%' OR historico_traduzido ILIKE '%cobran%' "
    "THEN 'Cobranca' ELSE 'Consultoria' END"
)
_FILTER = (
    "descricao ILIKE '%consultor%' OR historico_traduzido ILIKE '%consultor%' "
    "OR descricao ILIKE '%cobran%' OR historico_traduzido ILIKE '%cobran%'"
)

SQL_PROVISAO = f"""
WITH cpr AS (
  SELECT {_CAT} AS categoria,
         (regexp_match(descricao,'pagamento\\s+(\\d{{2}})/(\\d{{2}})/(\\d{{2}})')) AS m,
         data_posicao, valor
  FROM wh_cpr_movimento
  WHERE {_FILTER}
)
SELECT categoria,
       to_char(min(data_posicao),'MM/YYYY') AS competencia,
       CASE WHEN m IS NOT NULL
            THEN to_date('20'||m[3]||'-'||m[2]||'-'||m[1],'YYYY-MM-DD') END AS pagamento_agendado,
       count(*) AS dias_acumulo,
       min(data_posicao) AS inicio, max(data_posicao) AS fim,
       abs(min(valor)) AS valor_provisionado
FROM cpr
GROUP BY categoria, m
ORDER BY categoria, min(data_posicao);
"""

SQL_PAGAMENTOS = f"""
SELECT data_liquidacao AS data,
       {_CAT} AS categoria,
       descricao, historico_traduzido,
       abs(saidas) AS pago, entradas AS estorno, abs(entradas+saidas) AS liquido
FROM wh_movimento_caixa
WHERE carteira_cliente_nome='REALINVEST FIDC' AND ({_FILTER})
ORDER BY data_liquidacao, categoria;
"""

SQL_CONCILIACAO = f"""
WITH cpr_batch AS (
  SELECT {_CAT} AS categoria,
         (regexp_match(descricao,'pagamento\\s+(\\d{{2}})/(\\d{{2}})/(\\d{{2}})')) AS m,
         valor, data_posicao
  FROM wh_cpr_movimento WHERE {_FILTER}
),
cpr_peak AS (
  SELECT categoria,
    CASE WHEN m IS NOT NULL
         THEN to_char(to_date('20'||m[3]||'-'||m[2]||'-'||m[1],'YYYY-MM-DD'),'YYYY-MM')
         ELSE to_char(min(data_posicao),'YYYY-MM') END AS mes,
    min(valor) AS pico
  FROM cpr_batch GROUP BY categoria, m
),
cpr_m AS (SELECT mes, categoria, abs(sum(pico)) AS provisionado FROM cpr_peak GROUP BY mes, categoria),
caixa_m AS (
  SELECT to_char(data_liquidacao,'YYYY-MM') AS mes, {_CAT} AS categoria,
         abs(sum(entradas+saidas)) AS pago
  FROM wh_movimento_caixa
  WHERE carteira_cliente_nome='REALINVEST FIDC' AND ({_FILTER})
  GROUP BY 1,2
)
SELECT COALESCE(c.mes,k.mes) AS mes, COALESCE(c.categoria,k.categoria) AS categoria,
       c.provisionado, k.pago,
       COALESCE(c.provisionado,0)-COALESCE(k.pago,0) AS diferenca
FROM cpr_m c FULL JOIN caixa_m k ON c.mes=k.mes AND c.categoria=k.categoria
ORDER BY mes, categoria;
"""


def _f(v: object) -> object:
    return float(v) if isinstance(v, Decimal) else v


def _header(ws, labels: list[str]) -> None:
    ws.append(labels)
    for c in ws[1]:
        c.fill = _HEAD_FILL
        c.font = _HEAD_FONT
        c.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"


def _widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


async def main() -> None:
    conn = await asyncpg.connect(_load_dsn())
    try:
        prov = await conn.fetch(SQL_PROVISAO)
        pags = await conn.fetch(SQL_PAGAMENTOS)
        conc = await conn.fetch(SQL_CONCILIACAO)
    finally:
        await conn.close()

    wb = Workbook()

    # ── Leia-me ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Leia-me"
    notas = [
        ("Conferencia — Consultoria e Cobranca", True),
        ("Fundo: REALINVEST FIDC  |  Gerado em: " + date.today().strftime("%d/%m/%Y"), False),
        ("", False),
        ("Fontes:", True),
        ("  Provisao  -> wh_cpr_movimento (CPR = Contas a Pagar e Receber)", False),
        ("  Pagamento -> wh_movimento_caixa (visao de caixa QiTech)", False),
        ("", False),
        ("Convencoes:", True),
        ("  Valores em R$, magnitude positiva (sao despesas/saidas).", False),
        ("  Provisao: mes de competencia em MM/AAAA; pagamento agendado em DD/MM/AAAA.", False),
        ("  Pagamento: data de liquidacao em DD/MM/AAAA.", False),
        ("", False),
        ("ATENCAO (regra do acumulado):", True),
        ("  wh_cpr_movimento.valor e SALDO ACUMULADO da provisao (snapshot diario", False),
        ("  que cresce ao longo do mes e zera no pagamento). Por isso a coluna", False),
        ("  'Valor provisionado' usa o PICO do lote (saldo do ultimo dia), NAO a", False),
        ("  soma dos dias. Cada lote e identificado pelo texto 'com pagamento DD/MM/AA'.", False),
        ("", False),
        ("Classificacao: historico/descricao com 'cobran' -> Cobranca; senao Consultoria.", False),
    ]
    for txt, bold in notas:
        ws.append([txt])
        if bold:
            ws[ws.max_row][0].font = Font(bold=True)
    ws.column_dimensions["A"].width = 90

    # ── Provisao ─────────────────────────────────────────────────────────
    ws = wb.create_sheet("Provisao (CPR)")
    _header(ws, ["Categoria", "Competencia (MM/AAAA)", "Pagamento agendado",
                 "Dias acumulo", "Inicio acumulo", "Fim acumulo", "Valor provisionado (pico)"])
    for r in prov:
        ws.append([r["categoria"], r["competencia"], r["pagamento_agendado"],
                   r["dias_acumulo"], r["inicio"], r["fim"], _f(r["valor_provisionado"])])
        row = ws.max_row
        for col in (3, 5, 6):
            ws.cell(row, col).number_format = _DATE
        ws.cell(row, 7).number_format = _MONEY
    _total_row(ws, label_col=1, label="TOTAL", value_cols={7: sum(_f(r["valor_provisionado"]) for r in prov)})
    _widths(ws, [14, 20, 20, 13, 16, 16, 24])

    # ── Pagamentos ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Pagamentos (Caixa)")
    _header(ws, ["Data", "Categoria", "Rubrica (descricao)", "Historico traduzido",
                 "Pago", "Estorno", "Liquido"])
    for r in pags:
        ws.append([r["data"], r["categoria"], r["descricao"], r["historico_traduzido"],
                   _f(r["pago"]), _f(r["estorno"]), _f(r["liquido"])])
        row = ws.max_row
        ws.cell(row, 1).number_format = _DATE
        for col in (5, 6, 7):
            ws.cell(row, col).number_format = _MONEY
    _total_row(ws, label_col=1, label="TOTAL", value_cols={
        5: sum(_f(r["pago"]) for r in pags),
        6: sum(_f(r["estorno"]) for r in pags),
        7: sum(_f(r["liquido"]) for r in pags),
    })
    _widths(ws, [12, 14, 38, 30, 14, 14, 14])

    # ── Conciliacao ──────────────────────────────────────────────────────
    ws = wb.create_sheet("Conciliacao mensal")
    _header(ws, ["Mes pagamento (MM/AAAA)", "Categoria", "Provisionado", "Pago", "Diferenca"])
    for r in conc:
        mes = r["mes"]
        mes_fmt = f"{mes[5:7]}/{mes[0:4]}" if mes else None
        ws.append([mes_fmt, r["categoria"], _f(r["provisionado"]), _f(r["pago"]), _f(r["diferenca"])])
        row = ws.max_row
        for col in (3, 4, 5):
            ws.cell(row, col).number_format = _MONEY
    _widths(ws, [22, 14, 16, 16, 16])

    wb.save(OUT_PATH)
    print(f"OK -> {OUT_PATH}")
    print(f"  Provisao: {len(prov)} lotes | Pagamentos: {len(pags)} | Conciliacao: {len(conc)} linhas")


def _total_row(ws, *, label_col: int, label: str, value_cols: dict[int, float]) -> None:
    # +2 deixa 1 linha em branco de espacador. (append([]) nao avanca max_row
    # porque linha sem celula nao conta — escrever direto na posicao e seguro.)
    row = ws.max_row + 2
    ws.cell(row, label_col, label).font = Font(bold=True)
    for col, val in value_cols.items():
        c = ws.cell(row, col, val)
        c.font = Font(bold=True)
        c.number_format = _MONEY


if __name__ == "__main__":
    asyncio.run(main())
