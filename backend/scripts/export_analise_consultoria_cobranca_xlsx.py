"""ANALISE COMPLETA — Consultoria/Cobranca do REALINVEST FIDC (com PROVISAO ESTIMADA).

Reconciliacao PROVISAO (CPR, estimada) x PAGAMENTO (extrato, por empresa).

== PROVISAO ESTIMADA (importante) ==
A linha de provisao no CPR ('Despesa de Consultoria Especializada' / 'Despesa de
Servicos de Cobranca') ACUMULA diariamente (pro-rata) rumo ao valor de fechamento,
mas DEIXA DE SER SNAPSHOTADA alguns dias antes do fim do mes (some do relatorio).
O pico bruto (ultimo saldo captado) portanto SUBESTIMA a provisao real.

Correcao -> PROJECAO pelo ritmo recente:
  provisao_estimada = ultimo_saldo_captado + (ultimo_incremento_diario
                       x dias_uteis_restantes_ate_o_fim_do_mes)
Dias uteis vem de wh_dim_dia_util. Usa-se o ULTIMO incremento (ritmo recente),
NUNCA a media — porque a taxa pode mudar no meio do mes.

  *** MAIO/2026: a taxa de CONSULTORIA DOBROU a partir de 18/05 (2.500 -> 5.000/dia);
      cobranca seguiu 2.500/dia. Projecao usa o ritmo novo -> consultoria ~75.000,
      cobranca ~50.000, total ESTIMADO ~125.000 (acima dos ~100k dos meses anteriores). ***

TODOS os valores de provisao nesta planilha sao ESTIMATIVAS (projecao ao fim do mes).

Fontes:
  - PROVISAO  -> wh_cpr_movimento (saldo acumulado diario; lote via 'pagamento DD/MM/YY')
  - PAGAMENTO -> wh_extrato_bancario (contas conciliacao 4532551 + movimento 4532543)
  - DIAS UTEIS -> wh_dim_dia_util

Empresas (handoff mar/2025): ONBOARD (45934845000192) consultoria+cobranca 03/2025->;
  BLUESTONE (11314857000100) consultoria 03/2022->03/2025.

Le DATABASE_URL de backend/.env.
"""

from __future__ import annotations

import asyncio
import re
from collections import defaultdict
from datetime import date
from decimal import Decimal
from pathlib import Path

import asyncpg
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

BACKEND_DIR = Path(__file__).resolve().parent.parent
ENV_PATH = BACKEND_DIR / ".env"
OUT_PATH = Path(r"C:\app_gr\analise_consultoria_cobranca_estimada.xlsx")

_MONEY = "#,##0.00"
_DATE = "DD/MM/YYYY"
_HEAD_FILL = PatternFill("solid", fgColor="1F2937")
_HEAD_FONT = Font(bold=True, color="FFFFFF")
_WARN_FILL = PatternFill("solid", fgColor="FEF3C7")
_TENANT = "7f00cc2b-8bb4-483f-87b7-b1db24d20902"

_ONBOARD = "45934845000192"
_BLUESTONE = "11314857000100"
_VENDORS: dict[str, tuple[str, str, str]] = {
    _ONBOARD: ("Onboard", "ONBOARD CONSULTORIA LTDA", "Consultoria + Cobranca"),
    _BLUESTONE: ("Bluestone", "BLUESTONE CONSULTORIA E ASSESSORIA FINANCEIRA LTDA", "Consultoria"),
}
_CONTA_LABEL = {"4532551": "Conciliacao", "4532543": "Movimento"}
_REGIME = [(2025, m) for m in range(7, 13)] + [(2026, m) for m in range(1, 6)]


def _load_dsn() -> str:
    text = ENV_PATH.read_text(encoding="utf-8")
    m = re.search(r"^\s*DATABASE_URL\s*=\s*(.+)\s*$", text, re.MULTILINE)
    if not m:
        raise RuntimeError("DATABASE_URL not found in .env")
    return m.group(1).strip().strip('"').strip("'").replace("postgresql+asyncpg://", "postgresql://")


SQL_CPR = r"""
SELECT CASE WHEN descricao ILIKE '%cobran%' OR historico_traduzido ILIKE '%cobran%'
            THEN 'Cobranca' ELSE 'Consultoria' END AS categoria,
       (regexp_match(descricao,'pagamento\s+(\d{2})/(\d{2})/(\d{2})')) AS m,
       data_posicao, valor
FROM wh_cpr_movimento
WHERE descricao ILIKE '%consultor%' OR historico_traduzido ILIKE '%consultor%'
   OR descricao ILIKE '%cobran%' OR historico_traduzido ILIKE '%cobran%'
ORDER BY data_posicao;
"""

SQL_PAG = """
SELECT COALESCE(data_movimento, data_lancamento) AS data,
       contrapartida_doc AS cnpj, conta, valor, descricao
FROM wh_extrato_bancario
WHERE tipo = 'D' AND contrapartida_doc = ANY($1::text[])
ORDER BY COALESCE(data_movimento, data_lancamento), conta;
"""

SQL_CAL = """
SELECT data, eh_dia_util
FROM wh_dim_dia_util
WHERE tenant_id = $1 AND data BETWEEN '2025-07-01' AND '2026-06-30'
ORDER BY data;
"""


def _f(v: object) -> float:
    return float(v) if isinstance(v, Decimal) else float(v)


def _ym_label(ym: tuple[int, int]) -> str:
    return f"{ym[1]:02d}/{ym[0]}"


def _header(ws, labels: list[str]) -> None:
    ws.append(labels)
    for c in ws[1]:
        c.fill = _HEAD_FILL
        c.font = _HEAD_FONT
        c.alignment = Alignment(horizontal="center", wrap_text=True)
    ws.freeze_panes = "A2"


def _widths(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _bold(ws, cells: list, money: tuple[int, ...] = ()) -> None:
    ws.append(cells)
    for c in ws[ws.max_row]:
        c.font = Font(bold=True)
    for col in money:
        ws.cell(ws.max_row, col).number_format = _MONEY


async def main() -> None:
    conn = await asyncpg.connect(_load_dsn())
    try:
        cpr = await conn.fetch(SQL_CPR)
        pags = await conn.fetch(SQL_PAG, list(_VENDORS.keys()))
        cal = await conn.fetch(SQL_CAL, _TENANT)
    finally:
        await conn.close()

    # ── calendario: dias uteis + ultimo dia util por mes ─────────────────
    busdays = sorted(r["data"] for r in cal if r["eh_dia_util"])
    last_bd: dict[tuple[int, int], date] = {}
    for d in busdays:
        last_bd[(d.year, d.month)] = d  # ascendente -> fica o maior

    def remaining_bd(last_date: date, ym: tuple[int, int]) -> int:
        lbd = last_bd.get(ym)
        if lbd is None:
            return 0
        return sum(1 for d in busdays if last_date < d <= lbd)

    # ── monta lotes e projeta ────────────────────────────────────────────
    lotes_rows: dict[tuple, list[tuple[date, float]]] = defaultdict(list)
    for r in cpr:
        cat = r["categoria"]
        m = r["m"]
        if m:
            sched = date(2000 + int(m[2]), int(m[1]), int(m[0]))
            key = (cat, sched.isoformat())
        else:
            key = (cat, f"avulso-{r['data_posicao'].isoformat()}")
        lotes_rows[key].append((r["data_posicao"], _f(r["valor"])))

    # por lote: competencia, dias, saldo captado, ritmo recente, dias proj, estimado
    lotes = []
    for (cat, kid), rows in lotes_rows.items():
        rows.sort()
        dates = [d for d, _ in rows]
        vals = [v for _, v in rows]
        comp_ym = (dates[0].year, dates[0].month)
        last_date = dates[-1]
        captured = abs(vals[-1])
        avulso = kid.startswith("avulso-")
        sched_label = "(sem data)" if avulso else date.fromisoformat(kid).strftime("%d/%m/%Y")
        rate = abs(vals[-1] - vals[-2]) if len(vals) >= 2 else 0.0
        rem = 0 if avulso else remaining_bd(last_date, comp_ym)
        projected = captured + rate * rem
        lotes.append({
            "cat": cat, "comp": comp_ym, "ini": dates[0], "fim": last_date,
            "sched": sched_label,
            "ndias": len(rows), "captured": captured, "rate": rate,
            "rem": rem, "projected": projected, "avulso": avulso,
        })

    # agrega projetado por (competencia, categoria)
    prov_est: dict[tuple[int, int], dict] = {}
    for lo in lotes:
        b = prov_est.setdefault(lo["comp"], {"Consultoria": 0.0, "Cobranca": 0.0, "avulso": False})
        b[lo["cat"]] += lo["projected"]
        if lo["avulso"]:
            b["avulso"] = True

    # ── pagamentos normalizados ──────────────────────────────────────────
    pay = []
    for r in pags:
        c = r["cnpj"]
        curto, completo, servico = _VENDORS[c]
        pay.append({"data": r["data"], "cnpj": c, "curto": curto, "completo": completo,
                    "servico": servico, "conta": _CONTA_LABEL.get(r["conta"], r["conta"]),
                    "valor": _f(r["valor"]), "desc": r["descricao"]})
    onboard = [p for p in pay if p["cnpj"] == _ONBOARD]

    def _window(y: int, m: int) -> tuple[date, date]:
        return date(y, m, 20), (date(y + 1, 1, 10) if m == 12 else date(y, m + 1, 10))

    wb = Workbook()

    # ── Leia-me ──────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Leia-me"
    for txt, bold in [
        ("Analise — Consultoria / Cobranca (provisao ESTIMADA x pagamento)", True),
        ("Fundo: REALINVEST FIDC  |  Gerado em: " + date.today().strftime("%d/%m/%Y"), False),
        ("", False),
        (">>> TODOS os valores de PROVISAO sao ESTIMATIVAS (projecao ao fim do mes). <<<", True),
        ("", False),
        ("POR QUE ESTIMATIVA", True),
        ("  A linha de provisao acumula diariamente, mas some do relatorio alguns dias", False),
        ("  antes do fechamento. O ultimo saldo captado subestima a provisao real.", False),
        ("  Projetamos ate o ultimo dia util do mes:", False),
        ("    provisao_estimada = ultimo_saldo + (ritmo_recente_por_dia x dias_uteis_restantes)", False),
        ("  Ritmo = ULTIMO incremento diario (nao a media), pois a taxa pode mudar no mes.", False),
        ("  Dias uteis: wh_dim_dia_util. Ex. marco: 40.909 + 2.272,73 x 4 = 50.000 por categoria.", False),
        ("", False),
        ("COMPETENCIA = MES DE ACUMULO (regime de competencia)", True),
        ("  Cada lote e alocado no mes em que ACUMULOU (mes dos snapshots diarios), NAO", False),
        ("  na data 'pagamento DD/MM/YY' escrita na descricao (essa e so a data agendada).", False),
        ("  Ex.: o lote 'pagamento 08/04/26' acumula 02/03->25/03 -> alocado em MARCO/2026.", False),
        ("  A aba 'Provisao CPR (est.)' mostra as duas colunas lado a lado p/ conferencia.", False),
        ("", False),
        ("*** MAIO/2026 — ATENCAO ***", True),
        ("  A taxa de CONSULTORIA DOBROU a partir de 18/05 (2.500 -> 5.000/dia).", False),
        ("  Cobranca seguiu 2.500/dia. Projecao usa o ritmo novo:", False),
        ("    Consultoria ~75.000 + Cobranca ~50.000 = ESTIMADO ~125.000 (acima dos ~100k).", False),
        ("  Maio ainda esta acumulando — o numero firma no fechamento.", False),
        ("", False),
        ("EMPRESAS (handoff mar/2025)", True),
        ("  ONBOARD CONSULTORIA LTDA  (45934845000192) consultoria + cobranca, 03/2025->", False),
        ("  BLUESTONE CONS. E ASSESS. (11314857000100) consultoria, 03/2022->03/2025", False),
        ("", False),
        ("OUTROS ACHADOS", True),
        ("  - Provisao CPR comecou jul/2025; pagamentos anteriores (~R$ 620k) sem provisao.", False),
        ("  - Cobranca ficou BUNDLED na linha de consultoria ate jan/2026; split em fev/2026.", False),
        ("  - Ago/2025: ha um lancamento avulso de 50k (29/08, 1 dia, rotulo 'Consultoria')", False),
        ("    que infla a competencia — flag 'avulso' na aba Provisao CPR.", False),
        ("  - Pagamento atribuido a competencia via janela [dia 20 .. dia 10 do mes seguinte].", False),
    ]:
        ws.append([txt])
        if bold:
            ws[ws.max_row][0].font = Font(bold=True)
    ws.column_dimensions["A"].width = 98

    # ── Reconciliacao ────────────────────────────────────────────────────
    ws = wb.create_sheet("Reconciliacao")
    _header(ws, ["Competencia", "Consultoria (est.)", "Cobranca (est.)", "Total provisao (est.)",
                 "Pago (liquidacao)", "Delta (est.-pago)", "Delta acum.", "Observacao"])
    acc = 0.0
    tc = tcb = tp = 0.0
    for (y, m) in _REGIME:
        b = prov_est.get((y, m), {"Consultoria": 0.0, "Cobranca": 0.0, "avulso": False})
        cons, cobr = b["Consultoria"], b["Cobranca"]
        ptot = cons + cobr
        s, e = _window(y, m)
        pago = sum(p["valor"] for p in onboard if s <= p["data"] <= e)
        delta = ptot - pago
        acc += delta
        tc += cons; tcb += cobr; tp += pago
        obs = []
        if cobr == 0 and ptot > 0:
            obs.append("cobranca BUNDLED na consultoria")
        if b["avulso"]:
            obs.append("inclui lancamento avulso 50k (29/08)")
        if (y, m) == (2026, 5):
            obs.append("consultoria DOBROU em 18/05; maio ainda acumulando")
        if pago == 0:
            obs.append("nao liquidado ainda")
        ws.append([_ym_label((y, m)), cons, cobr or None, ptot, pago or None, delta, acc, "; ".join(obs)])
        for col in (2, 3, 4, 5, 6, 7):
            ws.cell(ws.max_row, col).number_format = _MONEY
        if b["avulso"] or (y, m) == (2026, 5):
            for c in ws[ws.max_row]:
                c.fill = _WARN_FILL
    _bold(ws, ["TOTAL regime (est.)", tc, tcb, tc + tcb, tp, (tc + tcb) - tp, "", ""], money=(2, 3, 4, 5, 6))
    ws.append([])
    ws.append(["PRE-REGIME (pago sem provisao CPR):"])
    ws[ws.max_row][0].font = Font(bold=True)
    blue = sum(p["valor"] for p in pay if p["cnpj"] == _BLUESTONE)
    onb_pre = sum(p["valor"] for p in onboard if p["data"] < date(2025, 7, 20))
    for lbl, val in [("  Bluestone (consultoria, 03/2022->03/2025)", blue),
                     ("  Onboard (inicial, 03/2025->06/2025)", onb_pre)]:
        ws.append([lbl, "", "", "", val, "", "", "sem provisao CPR"])
        ws.cell(ws.max_row, 5).number_format = _MONEY
    _widths(ws, [13, 16, 14, 16, 16, 16, 14, 46])

    # ── Provisao CPR (transparencia da projecao) ─────────────────────────
    ws = wb.create_sheet("Provisao CPR (est.)")
    _header(ws, ["Categoria", "Competencia (mes de acumulo)", "Pagamento agendado (texto)",
                 "Periodo captado", "Dias captados", "Saldo captado", "Ritmo recente (R$/dia)",
                 "Dias projetados", "Provisao estimada", "Obs"])
    for lo in sorted(lotes, key=lambda x: (x["cat"], x["comp"])):
        obs = "lancamento avulso (1 dia, sem projecao)" if lo["avulso"] else (
            f"+{lo['rem']}d x {lo['rate']:,.2f} projetado" if lo["rem"] else "captado ate o fim (sem projecao)")
        ws.append([lo["cat"], _ym_label(lo["comp"]), lo["sched"],
                   f"{lo['ini'].strftime('%d/%m/%y')} a {lo['fim'].strftime('%d/%m/%y')}",
                   lo["ndias"], lo["captured"], lo["rate"], lo["rem"], lo["projected"], obs])
        for col in (6, 7, 9):
            ws.cell(ws.max_row, col).number_format = _MONEY
        if lo["avulso"] or lo["rem"] >= 3:
            for c in ws[ws.max_row]:
                c.fill = _WARN_FILL
    _bold(ws, ["TOTAL", "", "", "", "", "", "", "", sum(lo["projected"] for lo in lotes), ""], money=(9,))
    _widths(ws, [12, 16, 16, 20, 9, 15, 16, 10, 16, 30])

    # ── Pagamentos ───────────────────────────────────────────────────────
    ws = wb.create_sheet("Pagamentos")
    _header(ws, ["Data", "Nome", "Empresa", "CNPJ", "Conta", "Valor", "Historico banco"])
    for p in pay:
        ws.append([p["data"], p["curto"], p["completo"], p["cnpj"], p["conta"], p["valor"], p["desc"]])
        ws.cell(ws.max_row, 1).number_format = _DATE
        ws.cell(ws.max_row, 6).number_format = _MONEY
    _bold(ws, ["TOTAL", "", "", "", "", sum(p["valor"] for p in pay), ""], money=(6,))
    _widths(ws, [12, 11, 46, 16, 14, 15, 46])

    # ── Resumo por empresa ───────────────────────────────────────────────
    ws = wb.create_sheet("Resumo por empresa")
    _header(ws, ["Empresa", "CNPJ", "Tipo de servico", "TEDs", "Total pago", "Primeiro", "Ultimo"])
    by_v: dict[str, dict] = {}
    for p in pay:
        v = by_v.setdefault(p["cnpj"], {"nome": p["completo"], "servico": p["servico"],
                                        "teds": 0, "total": 0.0, "min": p["data"], "max": p["data"]})
        v["teds"] += 1; v["total"] += p["valor"]
        v["min"] = min(v["min"], p["data"]); v["max"] = max(v["max"], p["data"])
    for c, v in sorted(by_v.items(), key=lambda kv: -kv[1]["total"]):
        ws.append([v["nome"], c, v["servico"], v["teds"], v["total"], v["min"], v["max"]])
        ws.cell(ws.max_row, 5).number_format = _MONEY
        ws.cell(ws.max_row, 6).number_format = _DATE
        ws.cell(ws.max_row, 7).number_format = _DATE
    _bold(ws, ["TOTAL", "", "", sum(v["teds"] for v in by_v.values()),
               sum(v["total"] for v in by_v.values()), "", ""], money=(5,))
    _widths(ws, [46, 16, 22, 8, 16, 14, 14])

    wb.save(OUT_PATH)
    print(f"OK -> {OUT_PATH}")
    print(f"  Lotes: {len(lotes)} | Pagamentos: {len(pay)}")
    print(f"  Provisao ESTIMADA regime: R$ {tc + tcb:,.2f} (cons {tc:,.2f} + cobr {tcb:,.2f})")
    print(f"  Pago regime: R$ {tp:,.2f}")
    for (y, m) in _REGIME:
        b = prov_est.get((y, m), {"Consultoria": 0.0, "Cobranca": 0.0})
        print(f"    {m:02d}/{y}: cons {b['Consultoria']:>10,.2f}  cobr {b['Cobranca']:>10,.2f}  "
              f"total {b['Consultoria'] + b['Cobranca']:>10,.2f}")


if __name__ == "__main__":
    asyncio.run(main())
